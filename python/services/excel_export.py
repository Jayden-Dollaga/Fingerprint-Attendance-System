###############################################################################
#  services/excel_export.py
#  AS608 Fingerprint Attendance System
#
#  Exports attendance records to Excel (.xlsx)
#  Uses openpyxl. Install: pip install openpyxl
###############################################################################

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ModuleNotFoundError:  # pragma: no cover
    openpyxl = None
    Font = PatternFill = Alignment = None

from core.database import get_attendance_today, get_attendance_all, get_attendance_by_date
from core.utils import get_export_path, timestamp_filename, today_str


HEADERS = ["ID", "Fingerprint ID", "Student No.", "Name", "Grade",
           "Section", "Date", "Time", "Confidence", "Status"]

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
ALT_FILL     = PatternFill("solid", fgColor="EBF0FA")


def _style_sheet(ws):
    """Apply header styling and column widths."""
    col_widths = [6, 14, 14, 20, 8, 12, 12, 10, 12, 14]
    for i, (cell, width) in enumerate(zip(ws[1], col_widths), 1):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill


def _write_rows(ws, records):
    """Write attendance records to worksheet."""
    for r in records:
        ws.append([
            r["id"],
            r["fingerprint_id"],
            r["student_no"],
            r["student_name"],
            r["grade"],
            r["section"],
            r["date"],
            r["time"],
            r["confidence"],
            r["status"],
        ])


def export_today():
    """
    Export today's attendance to Excel.
    Returns the full file path of the saved file.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install -r requirements.txt")

    records  = get_attendance_today()
    filename = timestamp_filename(f"attendance_{today_str()}", "xlsx")
    filepath = get_export_path(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {today_str()}"
    ws.append(HEADERS)
    _write_rows(ws, records)
    _style_sheet(ws)
    wb.save(filepath)

    return filepath, len(records)


def export_all():
    """
    Export all attendance records to Excel.
    Returns the full file path of the saved file.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install -r requirements.txt")

    records  = get_attendance_all()
    filename = timestamp_filename("attendance_all", "xlsx")
    filepath = get_export_path(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Attendance"
    ws.append(HEADERS)
    _write_rows(ws, records)
    _style_sheet(ws)
    wb.save(filepath)

    return filepath, len(records)


def export_by_date(date_str):
    """
    Export attendance for a specific date (YYYY-MM-DD).
    Returns the full file path of the saved file.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install -r requirements.txt")

    records  = get_attendance_by_date(date_str)
    filename = timestamp_filename(f"attendance_{date_str}", "xlsx")
    filepath = get_export_path(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {date_str}"
    ws.append(HEADERS)
    _write_rows(ws, records)
    _style_sheet(ws)
    wb.save(filepath)

    return filepath, len(records)
